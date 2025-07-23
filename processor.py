
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def format_currency(value: float) -> str:
    """Форматирует число в валюту с рублями"""
    if pd.isna(value) or value == 0:
        return "-   ₽"
    return f"{value:,.2f}".replace(",", " ") + " ₽"

def format_currency_numeric(value: float) -> float:
    """Форматирует число для сохранения в Excel (числовое значение)"""
    if pd.isna(value) or value == 0:
        return 0
    return float(value)

def format_percentage(value: float) -> str:
    """Форматирует процент"""
    if pd.isna(value) or value == 0:
        return "0%"
    # Если целое число, показываем без десятичных
    if value == int(value):
        return f"{int(value)}%"
    else:
        # Оставляем как есть, но убираем нули в конце
        return f"{value:.10g}%"

def read_wb_report(file_path: str) -> pd.DataFrame:
    """Читает файл отчета Wildberries (xlsx или csv)"""
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path)
    elif file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        raise ValueError("Файл должен быть в формате .xlsx или .csv")
    
    return df

def create_summary_data(df: pd.DataFrame) -> tuple:
    """Создает структурированные данные для отчета"""
    
    # Определение индексов столбцов
    col_indices = {}
    column_mapping = {
        'Тип документа': 'J',
        'Обоснование для оплаты': 'K', 
        'Кол-во': 'N',
        'Цена розничная': 'O',
        'К перечислению Продавцу за реализованный Товар': 'AH',
        'Удержания': 'BI',
        'Хранение': 'BH',
        'Услуги по доставке товара покупателю': 'AK',
        'Общая сумма штрафов': 'AO',
        'Эквайринг/Комиссии за организацию платежей': 'AC',
        'Платная приемка': 'BJ',
        'Номер поставки': 'B',
        'Возмещение издержек по перевозке/по складским операциям с товаром': 'BK',
        'Размер кВВ, %': 'X',
        'Виды логистики, штрафов и корректировок ВВ': 'AQ'  # Добавлен новый столбец
    }
    
    for col_name, col_letter in column_mapping.items():
        for idx, name in enumerate(df.columns):
            if name == col_name:
                col_indices[col_letter] = idx
                break
    
    # Фильтрация пустых строк
    df_filtered = df.dropna(subset=[df.columns[col_indices['J']], df.columns[col_indices['K']]], how='all')
    df_filtered = df_filtered.copy()
    df_filtered['Номер поставки'] = df_filtered.iloc[:, col_indices['B']]
    
    # Создание структурированных данных для первой таблицы
    structured_data = []
    
    # Категории для обработки
    categories = [
        {
            'name': 'Возврат',
            'type': 'both',  # Изменено на both - проверка и J и K
            'value_j': 'Возврат',
            'value_k': 'Возврат',
            'has_details': True
        },
        {
            'name': 'Логистика',
            'type': 'direct',
            'column': 'AK',
            'has_details': False
        },
        {
            'name': 'Продажа',
            'type': 'both',  # Новый тип - проверка и J и K
            'value_j': 'Продажа',
            'value_k': 'Продажа',
            'has_details': True
        },
        {
            'name': 'Возмещение издержек по перевозке/по складским операциям с товаром',
            'type': 'K',
            'value': 'Возмещение издержек по перевозке/по складским операциям с товаром',
            'has_details': False,
            'column': 'BK'
        },
        {
            'name': 'Хранение',
            'type': 'direct',
            'column': 'BH',
            'has_details': False
        },
        {
            'name': 'Удержание',
            'type': 'direct',
            'column': 'BI', 
            'has_details': False
        },
        {
            'name': 'Коррекция логистики',
            'type': 'K',
            'value': 'Коррекция логистики',
            'has_details': False,
            'column': 'AK'
        },
        {
            'name': 'Штраф',
            'type': 'direct',
            'column': 'AO',
            'has_details': False
        },
        {
            'name': 'Компенсация ущерба',
            'type': 'K',
            'value': 'Компенсация ущерба',
            'has_details': False
        },
        {
            'name': 'Добровольная компенсация при возврате',
            'type': 'K',
            'value': 'Добровольная компенсация при возврате',
            'has_details': False
        }
    ]
    
    # Сбор данных по категориям (только общие суммы, без детализации)
    category_totals = {}
    
    for category in categories:
        if category['type'] == 'J':
            filtered_data = df_filtered[df_filtered.iloc[:, col_indices['J']] == category['value']]
        elif category['type'] == 'K':
            filtered_data = df_filtered[df_filtered.iloc[:, col_indices['K']] == category['value']]
        elif category['type'] == 'both':  # Новая логика для возврата и продажи
            filtered_data = df_filtered[
                (df_filtered.iloc[:, col_indices['J']] == category['value_j']) & 
                (df_filtered.iloc[:, col_indices['K']] == category['value_k'])
            ]
        elif category['type'] == 'direct':
            filtered_data = df_filtered
            
        # Расчет значений для категории
        if category['name'] in ['Возврат', 'Продажа']:
            qty = filtered_data.iloc[:, col_indices['N']].sum()
            retail_price = (filtered_data.iloc[:, col_indices['N']] * filtered_data.iloc[:, col_indices['O']]).sum()
            to_seller = filtered_data.iloc[:, col_indices['AH']].sum()
            acquiring = filtered_data.iloc[:, col_indices['AC']].sum()
            retention = 0
            storage = 0
            logistics = 0
            fines = 0
            acceptance = 0
        elif category['name'] == 'Логистика':
            qty = 0
            retail_price = 0
            to_seller = 0
            acquiring = 0
            logistics = filtered_data.iloc[:, col_indices['AK']].sum()
            retention = 0
            storage = 0
            fines = 0
            acceptance = 0
        elif category['name'] == 'Хранение':
            qty = 0
            retail_price = 0
            to_seller = 0
            acquiring = 0
            storage = filtered_data.iloc[:, col_indices['BH']].sum()
            retention = 0
            logistics = 0
            fines = 0
            acceptance = 0
        elif category['name'] == 'Удержание':
            qty = 0
            retail_price = 0
            to_seller = 0
            acquiring = 0
            retention = filtered_data.iloc[:, col_indices['BI']].sum()
            storage = 0
            logistics = 0
            fines = 0
            acceptance = 0
        elif category['name'] == 'Коррекция логистики':
            qty = 0
            retail_price = 0
            to_seller = 0
            acquiring = 0
            logistics = filtered_data.iloc[:, col_indices['AK']].sum()
            retention = 0
            storage = 0
            fines = 0
            acceptance = 0
        elif category['name'] == 'Штраф':
            qty = 0
            retail_price = 0
            to_seller = 0
            acquiring = 0
            fines = filtered_data.iloc[:, col_indices['AO']].sum()
            retention = 0
            storage = 0
            logistics = 0
            acceptance = 0
        elif category['name'] == 'Возмещение издержек по перевозке/по складским операциям с товаром':
            qty = filtered_data.iloc[:, col_indices['N']].sum()
            retail_price = 0
            to_seller = 0
            acquiring = 0
            logistics = filtered_data.iloc[:, col_indices['BK']].sum()
            retention = 0
            storage = 0
            fines = 0
            acceptance = 0
        elif category['name'] in ['Компенсация ущерба', 'Добровольная компенсация при возврате']:
            qty = filtered_data.iloc[:, col_indices['N']].sum()
            retail_price = 0
            to_seller = filtered_data.iloc[:, col_indices['AH']].sum()
            acquiring = 0
            retention = 0
            storage = 0
            logistics = 0
            fines = 0
            acceptance = 0
        else:
            qty = 0
            retail_price = 0
            to_seller = 0
            acquiring = 0
            retention = 0
            storage = 0
            logistics = 0
            fines = 0
            acceptance = 0
            
        category_totals[category['name']] = {
            'qty': qty,
            'retail_price': retail_price,
            'to_seller': to_seller,
            'retention': retention,
            'storage': storage,
            'logistics': logistics,
            'fines': fines,
            'acceptance': acceptance,
            'acquiring': acquiring,
            'data': filtered_data,  # Сохраняем данные для детализации
            'raw_data': df_filtered  # Сохраняем все данные для второй таблицы
        }
    
    # Добавление строк категорий
    row_counter = 0
    for category in categories:
        category_data = category_totals[category['name']]
        
        # Добавление основной строки категории
        category_row = {
            'level': 0,
            'name': category['name'],
            'qty': category_data['qty'],
            'retail_price': category_data['retail_price'],
            'to_seller': category_data['to_seller'],
            'retention': category_data['retention'],
            'storage': category_data['storage'],
            'logistics': category_data['logistics'],
            'fines': category_data['fines'],
            'acceptance': category_data['acceptance'],
            'acquiring': category_data['acquiring'],
            'row_number': row_counter,
            'has_children': category.get('has_details', False) and len(category_data['data']) > 0
        }
        
        structured_data.append(category_row)
        row_counter += 1
        
        # Добавление детализации по номерам поставок
        if category.get('has_details', False) and len(category_data['data']) > 0:
            grouped = category_data['data'].groupby('Номер поставки').agg({
                df.columns[col_indices['N']]: 'sum',
                df.columns[col_indices['O']]: 'mean',
                df.columns[col_indices['AH']]: 'sum',
                df.columns[col_indices['AC']]: 'sum'
            }).reset_index()
            
            for _, row in grouped.iterrows():
                article_qty = row[df.columns[col_indices['N']]]
                article_retail = article_qty * row[df.columns[col_indices['O']]] if not pd.isna(row[df.columns[col_indices['O']]]) else 0
                
                # Номер поставки в числовом формате без десятичных
                supply_number = int(row['Номер поставки']) if not pd.isna(row['Номер поставки']) else str(row['Номер поставки'])
                
                detail_row = {
                    'level': 1,
                    'name': supply_number,
                    'qty': article_qty,
                    'retail_price': article_retail,
                    'to_seller': row[df.columns[col_indices['AH']]],
                    'retention': 0,
                    'storage': 0,
                    'logistics': 0,
                    'fines': 0,
                    'acceptance': 0,
                    'acquiring': row[df.columns[col_indices['AC']]],
                    'row_number': row_counter,
                    'has_children': False
                }
                
                structured_data.append(detail_row)
                row_counter += 1
    
    # Расчет общего итога (только по основным категориям)
    total_qty = sum(category_totals[cat['name']]['qty'] for cat in categories if cat['name'] in category_totals)
    total_retail = sum(category_totals[cat['name']]['retail_price'] for cat in categories if cat['name'] in category_totals)
    total_seller = sum(category_totals[cat['name']]['to_seller'] for cat in categories if cat['name'] in category_totals)
    total_retention = sum(category_totals[cat['name']]['retention'] for cat in categories if cat['name'] in category_totals)
    total_storage = sum(category_totals[cat['name']]['storage'] for cat in categories if cat['name'] in category_totals)
    total_logistics = sum(category_totals[cat['name']]['logistics'] for cat in categories if cat['name'] in category_totals)
    total_fines = sum(category_totals[cat['name']]['fines'] for cat in categories if cat['name'] in category_totals)
    total_acceptance = sum(category_totals[cat['name']]['acceptance'] for cat in categories if cat['name'] in category_totals)
    total_acquiring = sum(category_totals[cat['name']]['acquiring'] for cat in categories if cat['name'] in category_totals)
    
    # Добавление строки "Общий итог"
    total_row = {
        'level': 0,
        'name': 'Общий итог',
        'qty': total_qty,
        'retail_price': total_retail,
        'to_seller': total_seller,
        'retention': total_retention,
        'storage': total_storage,
        'logistics': total_logistics,
        'fines': total_fines,
        'acceptance': total_acceptance,
        'acquiring': total_acquiring,
        'row_number': row_counter,
        'has_children': False,
        'is_total': True
    }
    structured_data.append(total_row)
    
    # Создание данных для второй таблицы
    second_table_data = create_second_table_data(df_filtered, col_indices, category_totals)
    
    return structured_data, second_table_data

def create_second_table_data(df_filtered, col_indices, category_totals):
    """Создает данные для второй таблицы"""
    
    # Получаем данные для разных категорий
    sales_data = category_totals.get('Продажа', {})
    returns_data = category_totals.get('Возврат', {})
    voluntary_compensation_data = category_totals.get('Добровольная компенсация при возврате', {})
    compensation_damage_data = category_totals.get('Компенсация ущерба', {})
    
    # Продажи GROSS (розничная цена из продаж)
    sales_gross = sales_data.get('retail_price', 0)
    
    # ВБ компенсирует ущерб (сумма к перечислению из добровольной компенсации)
    vb_compensates_damage = voluntary_compensation_data.get('to_seller', 0)
    
    # Процент с продаж Wildberries (N * O * X/100) - теперь делим на 100
    sales_rows = df_filtered[
        (df_filtered.iloc[:, col_indices['J']] == 'Продажа') & 
        (df_filtered.iloc[:, col_indices['K']] == 'Продажа')
    ]
    wb_commission = (sales_rows.iloc[:, col_indices['N']] * 
                     sales_rows.iloc[:, col_indices['O']] * 
                     sales_rows.iloc[:, col_indices['X']] / 100).sum()  # Делим на 100
    
    # Эквайринг (из продаж)
    acquiring = sales_data.get('acquiring', 0)
    
    # Возвраты заказов (сумма к перечислению из возвратов)
    returns_amount = returns_data.get('to_seller', 0)
    
    # Логистика
    logistics = category_totals.get('Логистика', {}).get('logistics', 0)
    logistics += category_totals.get('Коррекция логистики', {}).get('logistics', 0)
    logistics += category_totals.get('Возмещение издержек по перевозке/по складским операциям с товаром', {}).get('logistics', 0)
    
    # Реклама (строки с "Оказание услуг «ВБ.Продвижение»" в AQ, значение из BI)
    advertising_rows = df_filtered[df_filtered.iloc[:, col_indices['AQ']] == 'Оказание услуг «ВБ.Продвижение»']
    advertising = advertising_rows.iloc[:, col_indices['BI']].sum()
    
    # Хранение
    storage = category_totals.get('Хранение', {}).get('storage', 0)
    
    # Штрафы
    fines = category_totals.get('Штраф', {}).get('fines', 0)
    
    # Платная приемка (сумма по всем строкам столбца BJ)
    paid_acceptance = df_filtered.iloc[:, col_indices['BJ']].sum()
    
    # Удержание (строки с "Удержание" в K, но не "Оказание услуг «ВБ.Продвижение»" в AQ)
    retention_rows = df_filtered[
        (df_filtered.iloc[:, col_indices['K']] == 'Удержание') & 
        (df_filtered.iloc[:, col_indices['AQ']] != 'Оказание услуг «ВБ.Продвижение»')
    ]
    retention = retention_rows.iloc[:, col_indices['BI']].sum()
    
    # Компенсация ущерба
    compensation_damage = compensation_damage_data.get('to_seller', 0)
    
    # Налоги и себестоимость (оставляем как в примере - пустые значения)
    taxes = 0  # Нет данных
    cost_of_goods = 0  # Нет данных
    
    # Создание структуры второй таблицы
    second_table = [
        {'name': 'Продажи GROSS', 'amount': sales_gross, 'percent': 100 if sales_gross > 0 else 0},
        {'name': 'ВБ компенсирует ущерб', 'amount': vb_compensates_damage, 'percent': (vb_compensates_damage / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Процент с продаж вайлдберриз', 'amount': wb_commission, 'percent': (wb_commission / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Эквайринг', 'amount': acquiring, 'percent': (acquiring / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Возвраты заказов', 'amount': returns_amount, 'percent': (returns_amount / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Логистика', 'amount': logistics, 'percent': (logistics / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Реклама', 'amount': advertising, 'percent': (advertising / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Подписка "Джем"', 'amount': 0, 'percent': 0},
        {'name': 'Хранение', 'amount': storage, 'percent': (storage / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Штрафы', 'amount': fines, 'percent': (fines / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Платная приемка', 'amount': paid_acceptance, 'percent': (paid_acceptance / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Удержание', 'amount': retention, 'percent': (retention / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Услуги транзитных поставок', 'amount': 0, 'percent': 0},
        {'name': 'Компенсация ущерба', 'amount': compensation_damage, 'percent': (compensation_damage / sales_gross * 100) if sales_gross > 0 else 0},
        {'name': 'Налоги', 'amount': taxes, 'percent': 0},  # Нет данных
        {'name': 'Себестоимость продукта', 'amount': cost_of_goods, 'percent': 0}  # Нет данных
    ]
    
    # Расчет итога
    # Итог = Продажи GROSS + ВБ компенсирует ущерб - все остальные расходы
    income_items = [second_table[0]['amount'], second_table[1]['amount']]  # Продажи GROSS и ВБ компенсирует ущерб
    expense_items = [item['amount'] for item in second_table[2:-1]]  # Все остальные кроме итога
    
    total_amount = sum(income_items) - sum(expense_items)
    
    # Проценты: доходы - расходы
    income_percent = sum([second_table[0]['percent'], second_table[1]['percent']])
    expense_percent = sum([item['percent'] for item in second_table[2:-1]])
    total_percent = income_percent - expense_percent
    
    second_table.append({'name': 'Итого:', 'amount': total_amount, 'percent': total_percent})
    
    return second_table

def create_excel_with_grouping(structured_data_list, second_table_data_list, output_path: str):
    """Создает Excel файл с двумя листами и группировкой строк"""
    
    wb = Workbook()
    
    # Первый лист - основная таблица
    ws1 = wb.active
    ws1.title = "Основной отчет"
    
    # Заголовки
    headers = ['Названия строк', 'Кол-во продаж', 'Розничная Цена', 'Сумма к перечислению продавцу', 
               'Удержание', 'Хранение товара', 'Логистика', 'Штрафы', 'Приемка платная', 'Эквайринг']
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Добавление заголовков на первый лист
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Добавление данных на первый лист
    for i, item in enumerate(structured_data_list):
        row_num = i + 2  # +2 потому что первая строка - заголовки
        
        # Преобразование данных для Excel (числовые значения)
        qty = format_currency_numeric(item['qty'])
        retail_price = format_currency_numeric(item['retail_price'])
        to_seller = format_currency_numeric(item['to_seller'])
        retention = format_currency_numeric(item['retention'])
        storage = format_currency_numeric(item['storage'])
        logistics = format_currency_numeric(item['logistics'])
        fines = format_currency_numeric(item['fines'])
        acceptance = format_currency_numeric(item['acceptance'])
        acquiring = format_currency_numeric(item['acquiring'])
        
        # Добавление данных в строку
        name_cell = ws1.cell(row=row_num, column=1, value=item['name'])
        qty_cell = ws1.cell(row=row_num, column=2, value=qty if item['name'] != 'Общий итог' else item['qty'])
        retail_cell = ws1.cell(row=row_num, column=3, value=retail_price)
        seller_cell = ws1.cell(row=row_num, column=4, value=to_seller)
        retention_cell = ws1.cell(row=row_num, column=5, value=retention)
        storage_cell = ws1.cell(row=row_num, column=6, value=storage)
        logistics_cell = ws1.cell(row=row_num, column=7, value=logistics)
        fines_cell = ws1.cell(row=row_num, column=8, value=fines)
        acceptance_cell = ws1.cell(row=row_num, column=9, value=acceptance)
        acquiring_cell = ws1.cell(row=row_num, column=10, value=acquiring)
        
        # Применение стилей
        cells = [name_cell, qty_cell, retail_cell, seller_cell, retention_cell, 
                storage_cell, logistics_cell, fines_cell, acceptance_cell, acquiring_cell]
        
        for cell in cells:
            cell.border = border
            if item.get('is_total', False):
                cell.font = total_font
                cell.fill = total_fill
            elif item['level'] == 1:  # Детализация
                cell.alignment = Alignment(indent=2)  # Отступ для вложенных строк
        
        # Установка уровня группировки
        if item['level'] == 1:  # Детализация (дочерние строки)
            ws1.row_dimensions[row_num].outline_level = 1
            ws1.row_dimensions[row_num].hidden = True  # Скрыты по умолчанию
    
    # Форматирование числовых столбцов на первом листе
    for row in range(2, ws1.max_row + 1):
        # Кол-во продаж - числовой формат
        qty_cell = ws1.cell(row=row, column=2)
        if qty_cell.value is not None and qty_cell.value != 0:
            qty_cell.number_format = '#,##0'
        
        # Номер поставки - числовой формат без десятичных
        name_cell = ws1.cell(row=row, column=1)
        if len(structured_data_list) > row-2 and structured_data_list[row-2]['level'] == 1:  # Только для детализации
            name_cell.number_format = '0'
        
        # Денежные столбцы
        money_columns = [3, 4, 5, 6, 7, 8, 9, 10]  # Розничная Цена и далее
        for col in money_columns:
            cell = ws1.cell(row=row, column=col)
            if cell.value is not None and cell.value != 0:
                cell.number_format = '#,##0.00" ₽"'
    
    # Автоширина колонок на первом листе
    for col in range(1, 11):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws1.max_row + 1):
            cell_value = ws1.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 30)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Закрепление заголовков на первом листе
    ws1.freeze_panes = 'A2'
    
    # Второй лист - таблица Россия
    ws2 = wb.create_sheet("Россия")
    
    # Заголовки для второй таблицы
    ws2.cell(row=1, column=1, value="").font = header_font
    ws2.cell(row=1, column=1, value="").fill = header_fill
    ws2.cell(row=1, column=1, value="").border = border
    ws2.cell(row=1, column=2, value="").font = header_font
    ws2.cell(row=1, column=2, value="").fill = header_fill
    ws2.cell(row=1, column=2, value="").border = border
    ws2.cell(row=1, column=3, value="%").font = header_font
    ws2.cell(row=1, column=3, value="%").fill = header_fill
    ws2.cell(row=1, column=3, value="%").border = border
    
    # Добавление данных на второй лист
    for i, item in enumerate(second_table_data_list):
        row_num = i + 2
        name_cell = ws2.cell(row=row_num, column=1, value=item['name'])
        amount_cell = ws2.cell(row=row_num, column=2, value=item['amount'])
        percent_cell = ws2.cell(row=row_num, column=3, value=item['percent']/100)  # Делим на 100 для правильного отображения
        
        # Применение стилей
        name_cell.border = border
        amount_cell.border = border
        percent_cell.border = border
        
        if item['name'] == 'Итого:':
            name_cell.font = Font(bold=True)
            amount_cell.font = Font(bold=True)
            percent_cell.font = Font(bold=True)
        
        # Форматирование
        # Для налогов и себестоимости - пустые значения
        if item['name'] in ['Налоги', 'Себестоимость продукта']:
            amount_cell.value = "-   ₽"
            percent_cell.value = 0
            percent_cell.number_format = '0%'
        elif item['amount'] != 0:
            amount_cell.number_format = '#,##0.00" ₽"'
        else:
            amount_cell.value = "-   ₽"
            
        if item['percent'] != 0:
            percent_cell.number_format = '0.00%'  # Будет отображаться правильно
        else:
            percent_cell.number_format = '0%'
    
    # Автоширина колонок на втором листе
    for col in range(1, 4):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws2.max_row + 1):
            cell_value = ws2.cell(row=row, column=col).value
            if cell_value:
                if isinstance(cell_value, str):
                    max_length = max(max_length, len(cell_value))
                else:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 30)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"Файл сохранен: {output_path}")

def process_wb_report_file(file_path: str, output_path: str = "результат_с_группировкой.xlsx"):
    """
    Основная функция для обработки отчета Wildberries с группировкой
    
    Args:
        file_path (str): Путь к файлу отчета (.xlsx или .csv)
        output_path (str): Путь для сохранения результата с группировкой
    """
    # Чтение файла
    df = read_wb_report(file_path)
    
    # Создание структурированных данных
    structured_data, second_table_data = create_summary_data(df)
    
    # Создание Excel файла с группировкой
    create_excel_with_grouping(structured_data, second_table_data, output_path)
    
    return structured_data, second_table_data

# Пример использования:
if __name__ == "__main__":
    # Укажите путь к вашему файлу
    file_path = "Отчёт.xlsx"  # или .csv
    
    # Обработка файла
    result1, result2 = process_wb_report_file(file_path, "результат_с_группировкой.xlsx")
    
    print("Обработка завершена! Откройте Excel файл и используйте функцию группировкой для сворачивания/разворачивания строк.")
